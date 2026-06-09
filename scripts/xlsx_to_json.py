import json
import openpyxl
from pathlib import Path

ROOT = Path(__file__).parent.parent
xlsx_path = ROOT / "deals_data.xlsx"
json_path = ROOT / "src" / "data" / "deals.json"

wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

headers = [cell.value for cell in ws[1]]

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    record = {}
    for key, val in zip(headers, row):
        record[key] = "" if val is None else val
    rows.append(record)

json_path.parent.mkdir(parents=True, exist_ok=True)
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)

print(f"[xlsx_to_json] Wrote {len(rows)} rows -> src/data/deals.json")
for r in rows:
    print(f"  ID={r.get('ID')} | {r.get('Fabric')} | {r.get('Name')} | img={'YES' if r.get('Photo_Url') else 'BLANK'}")
