"""
Run this whenever you update deals_data.xlsx:
    python scripts/xlsx2json.py
It overwrites public/deals_data.json which the website reads.
"""
import openpyxl, json, pathlib, sys

ROOT   = pathlib.Path(__file__).parent.parent
SRC    = ROOT / "public" / "deals_data.xlsx"
DEST   = ROOT / "public" / "deals_data.json"

if not SRC.exists():
    sys.exit(f"ERROR: {SRC} not found")

wb = openpyxl.load_workbook(SRC)
ws = wb.active

headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None or v == '' for v in row):
        continue
    rows.append({h: (str(v).strip() if v is not None else '') for h, v in zip(headers, row)})

DEST.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"✓ Exported {len(rows)} rows → {DEST}")
